from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.views import View
from .s3_handler import S3Handler
from openpyxl import load_workbook, Workbook
import requests
import os
import time
import json
import random
from .models import mtd_tokens,posted ,testuser, apps_ids, log1
from urllib.parse import urlparse
import logging

# Create your views here.
class HomeView(View):
    def get(self, request):
        print('home')
        # return HttpResponse('homeee')
        
        return render(request, 'pages/index.html', {'full_path': request.get_full_path()})

class MtdView(View):
    template_name = 'pages/mtd/uploadfile.html'

    def get(self, request):
        s3_handler = S3Handler()
        file_list = s3_handler.get_file_list()
        return render(request, self.template_name,{'vat': file_list, 'full_path': request.get_full_path()})

class FileListView(View):
    
    def get(self, request):

       # if not request.user.is_authenticated:
        #    return JsonResponse(status=401)

        s3_handler = S3Handler()
        file_list = s3_handler.get_file_list()

        return JsonResponse(data={'data': file_list}, safe=False)

class UploadFileView(View):

    def post(self, request):
        file = request.FILES['file']  
        file_name = file.name.split(".")[0]+'_'+str(random.randint(9999,9999999))
        full_file_name = file_name+os.path.splitext(file.name)[1]
        
        # wb = load_workbook(filename=file)

        # file_path = "%s%s" % (time.time(), os.path.splitext(file.name)[1])

        # wb.save(file_path)

        wb = load_workbook(filename=file)
        file_path = "%s%s" % (time.time(), os.path.splitext(full_file_name)[1])
        wb.save(file_path)
        s3_handler = S3Handler()
        s3_handler.upload_file(file_path, full_file_name)
        os.remove(file_path)
        #updated to it goes to 'mtd' path which is the intended
        return redirect('/mtd/')


class FileEditView(View):
    template_name = 'pages/mtd/file_edit.html'
    
    def get(self, request):
            
        api_response_code = None
        api_response_message = None
        if request.GET.get('api_response_code'):
            api_response_code = request.GET.get('api_response_code')
        if request.GET.get('api_response_message'):
            api_response_message = request.GET.get('api_response_message')
        key = request.GET.get('key')
        s3_handler = S3Handler()
        download_path = s3_handler.download_file(key)
        wb = load_workbook(download_path)
        ws = wb[wb.sheetnames[0]]
        data = []
        for col in ws[1]:
            data.append({'header': col.value})

        for col in ws[2]:
            data[col.column - 1]['value'] = col.value
        os.remove(download_path)
        return render(request, self.template_name, {'data': data, 'key': key, 'api_response_code': api_response_code, 'api_response_message': api_response_message})

    

    def post(self, request):
        if 'save' in request.POST:

            key = request.POST.get('key')
            headers = ['periodKey', 'vatDueSales', 'vatDueAcquisitions', 'totalVatDue', 'vatReclaimedCurrPeriod',
                    'netVatDue', 'totalValueSalesExVAT', 'totalValuePurchasesExVAT', 'totalValueGoodsSuppliedExVAT',
                    'totalAcquisitionsExVAT', 'finalised']
            wb = Workbook()
            ws = wb.active
            ws.append(headers)
            data = []
            for header in headers:
                data.append(request.POST.get(header))
            ws.append(data)
            
            file_path = "%s%s" % (time.time(), os.path.splitext(key)[1])
            wb.save(file_path)
            
            s3_handler = S3Handler()
            s3_handler.upload_file(file_path, key)
            os.remove(file_path)
            return redirect('/mtd/')
        else:
            periodKey = request.POST.get('periodKey')
            vatDueSales = request.POST.get('vatDueSales')
            vatDueAcquisitions = request.POST.get('vatDueAcquisitions')
            totalVatDue = request.POST.get('totalVatDue')
            vatReclaimedCurrPeriod = request.POST.get('vatReclaimedCurrPeriod')
            netVatDue = request.POST.get('netVatDue')
            totalValueSalesExVAT = request.POST.get('totalValueSalesExVAT')
            totalValuePurchasesExVAT = request.POST.get('totalValuePurchasesExVAT')
            totalValueGoodsSuppliedExVAT = request.POST.get('totalValueGoodsSuppliedExVAT')
            totalAcquisitionsExVAT = request.POST.get('totalAcquisitionsExVAT')
            finalised = request.POST.get('finalised')
            
            #tok=mtd_tokens.objects.get(id=8)
            tok=mtd_tokens.objects.latest('id')
            print('tok')
            print(tok)
            print(periodKey)
            endpoint='https://test-api.service.hmrc.gov.uk/organisations/vat/'+tok.vrn+'/returns'
        

            #vrn='700695645'
            #endpoint='https://test-api.service.hmrc.gov.uk/organisations/vat/'+vrn+'/returns'
            post_Data3={
                "periodKey": periodKey,
                "vatDueSales": vatDueSales,
                "vatDueAcquisitions": vatDueAcquisitions,
                "totalVatDue": totalVatDue,
                "vatReclaimedCurrPeriod": vatReclaimedCurrPeriod,
                "netVatDue": netVatDue,
                "totalValueSalesExVAT": totalValueSalesExVAT,
                "totalValuePurchasesExVAT": totalValuePurchasesExVAT,
                "totalValueGoodsSuppliedExVAT": totalValueGoodsSuppliedExVAT,
                "totalAcquisitionsExVAT": totalAcquisitionsExVAT,
                "finalised": True
                }
            headers ={
                    #"Authorization": "Bearer 61d270d116111693fd3fa8e061b9753c",
                    "Authorization": "Bearer " + tok.access_token,
                    "Accept" :"application/vnd.hmrc.1.0+json",
                    "Content-Type":"application/json"
                #"Gov-Client-Connection-Method": "DESKTOP_APP_VIA_SERVER"
            }
            response_vat1 = requests.post(endpoint,data=json.dumps(post_Data3), headers=headers)
            key = request.POST.get('key')

            try:
                response_vat=response_vat1.json()
            #except:
            except (json.decoder.JSONDecodeError, KeyError, ValueError) as e: 

                response_vat ={'code': 'DEAD', 'message': 'Server Issue'}
                print ('response json error in PostVat')
                l1=log1(src_sys_id='PostVat2', status1=' fail: Json deode Error '+ str(e) , 
                status2= str(response_vat1.status_code) +response_vat1.reason)
                l1.save()
                
            else:
                if 'code' in response_vat:
                    #zz=response_vat["code"]
                    zz=response_vat
                    print('zz')
                    print(zz)
                    # print(zz['code'])
                    # yy=zz["errors"]
                    
                    l1=log1(src_sys_id='PostVat2', status1=' fail JSON ok but error +vrn=' + str(tok.vrn), 
                        status2=response_vat["code"] +' '+ zz['code']+'  '+ zz['message']  , 
                        status3=str(response_vat1.status_code) +response_vat1.reason +'post period' +post_Data3['periodKey'] )
                    l1.save()
                    return render (request, 'pages/mtd/mtd_home2.html',{ 
                         'vat': response_vat,
                         'comment': 'PostVat JSON error in response and code is:' + tok.vrn
                         })
                else: #assume we are posting as no json decode error or post json decode 'error'
                        #if response_vat1
                        t1=posted( periodKey=post_Data3['periodKey'],               vrn=tok.vrn,
                        vatDueSales=post_Data3['vatDueSales'],           
                           vatDueAcquisitions=post_Data3["vatDueAcquisitions"],
                        #totalVatDue=post_Data3["totalVatDue"],
                        vatReclaimedCurrPeriod=post_Data3["vatReclaimedCurrPeriod"],    
                                    netVatDue=post_Data3["netVatDue"],
                        totalValueSalesExVAT=["totalValueSalesExVAT"],              totalValuePurchasesExVAT=post_Data3["totalValuePurchasesExVAT"],
                        totalValueGoodsSuppliedExVAT=["totalValueGoodsSuppliedExVAT"],
                                      totalAcquisitionsExVAT=["totalAcquisitionsExVAT"],
                        finalised=post_Data3["finalised"],
                                      access_token=response_vat["processingDate"],
                        endpoint=response_vat["paymentIndicator"],
                                      date=response_vat["formBundleNumber"]
                        #vrn=response_vat["chargeRefNumber"]

                        )
                        t1.save()
                    
                        l1=log1(src_sys_id='PostVat2', status1=' else:success ',
                            status3=str(response_vat1.status_code) +response_vat1.reason,
                            status2=tok.vrn)
                        l1.save()
                        #lj take this out - no 'code' as this branch ot try/except/else has no error
                       #return redirect('file-edit?key='+key+'&api_response_code='+response_vat["code"]+'&api_response_message='+response_vat["message"]+'')
                        return redirect('/mtd/')

class TestView(View):
    def get(self, request):
        # tok=mtd_tokens.objects.latest('id')
        # tok = '99a47b444c236f4e673c19d3961f1d'
        # print('tok')
        # print(tok)
        # obj=apps_ids.objects.get(id=5)
        # posted_obj = posted.objects.latest('id') #filter(vrn=tok.vrn)[:0].get()
        # periodpj=posted_obj.periodKey
        periodpj="0001"
        tokpj='9b51aaa6f3cc176bff8917e8145d528'
    

        endpoint='https://test-api.service.hmrc.gov.uk/organisations/vat/'+'997382159' +'/returns/'+ periodpj #"%23001"  

        #endpoint='https://test-api.service.hmrc.gov.uk/organisations/vat/'+tok.vrn +'/obligations/'

        headers ={
        'Authorization': 'Bearer '+ tokpj,  #'5c5adcd9d4ee25ac173c78225ef45462',
        'Accept' :'application/vnd.hmrc.1.0+json',
        #'Gov-Test-Scenario':'QUARTERLY_NONE_MET',
        #	'Gov-Test-Scenario':'INVALID_PERIODKEY',

        'Content-Type':'application/json'
        }


        post_Data2={
        'from': '2019-06-30', 
        'to': '2020-05-31',
        #'status':'F'	 
        }

        r= requests.get(endpoint,data=json.dumps(post_Data2), headers=headers)



        print(r.request.body)


        print(r.request.headers)

        print(r.request.url)
        print(r)
        print('abc')
        rj=r.json()

        print(rj)
        return HttpResponse('abc')


class InitView(View):
    def get(self, request):
        t= apps_ids(appname= 'bts1',		
                    client_Secrets='f18af97f-2acb-41f9-92d2-e373c44618a8',	
                    client_id='ZjScpZHoNA6TQJytfrEkFcSryf4a',
                    server_token='9d41fd904fc3e6449a51381f76e43df'
                    
                )
        t.save()
        t= apps_ids(appname= 'bts1test',        
                    client_Secrets='a97c7346-a779-403c-ba18-34abb71374c8',  
                    client_id='fQ84pQWgR71bsRWaHRSKRFsh5JQa',
                    server_token='98918245c5488ac09cf5a55955c5e1cd'
                    
                )
        t.save()
        parsed_uri = urlparse(request.build_absolute_uri())
        base_url = '{uri.scheme}://{uri.netloc}/'.format(uri=parsed_uri)+'mtd/mtd2/'
        
        if 'https' in base_url:
            mtd_url = base_url
        elif 'http' in base_url:
            mtd_url = base_url.replace("http","https")
        else:
            mtd_url = base_url
        
        return render(request, 'pages/mtd/apis/tu.html',{'mtd_url':mtd_url})

class sa_testuser(View):  #testuser organization code
    def get(self, request):
        obj=apps_ids.objects.get(id=10)
        parsed_uri = urlparse(request.build_absolute_uri())
        base_url = '{uri.scheme}://{uri.netloc}/'.format(uri=parsed_uri)+'mtd/mtd2agent/'  #this section allows code to be deployed to any hostname + uri
        
        if 'https' in base_url:
            mtd_url = base_url
        elif 'http' in base_url:
            mtd_url = base_url.replace("http","https")
        else:
            mtd_url = base_url
        endpoint = 'https://test-api.service.hmrc.gov.uk/create-test-user/individuals'        
        
        #endpoint = 'https://test-api.service.hmrc.gov.uk/create-test-user/organisations'		
        #endpoint='https://test-api.service.hmrc.gov.uk/create-test-user/agents'		

        post_Data2={
            "serviceNames":       ["mtd-income-tax"] # ["agent-services"] # ["mtd-vat"]
        }
        headers ={
            'Authorization': 'Bearer '+ obj.server_token,
            'Accept' :'application/vnd.hmrc.1.0+json',
            'Content-Type':'application/json',
            
        }

        r = requests.post(endpoint,data=json.dumps(post_Data2), headers=headers)
        #refactor using error codes per doc
        #try:
        #    response_vat=r.json()
        if r.status_code == 403 or r.status_code == 403 or r.status_code == 404 or r.status_code == 405 or r.status_code == 406 or r.status_code == 429 or r.status_code == 500 or r.status_code == 501 or r.status_code == 503 or r.status_code == 504:
            l1=log1(src_sys_id='Testuser_sa', status1='documented API error' , #+ str(e),
                status2=str(r.status_code) +' status reason'+ r.reason)
            l1.save()
            return render (request,  'pages/mtd/apis/tuagent_sa.html',
                    {#'vat': response_vat	,
                    'endpoint': endpoint,
                    'pd2':post_Data2,
                    'headers1': headers,
                    'title': 'testCreate_sa: FAIL' + str(r.status_code) + ' ' + str(r.reason),
                    'clientid': obj.client_id,
                    'mtd_url': mtd_url })	
        elif r.status_code == 201:
            response_vat=r.json()  #next refctor to rj
            t2=testuser(endpoint=endpoint,  #next step create agent table
                #vrn=response_vat['vrn'],
                vrn=response_vat['nino'],
                userid= response_vat['userId'],
                password=response_vat['password'],
                src_sys_id='create_tu_sa',
                appname=obj.appname 
                )
            t2.save()
            l1=log1(src_sys_id='create_testuser_sa', status1=' Try  else(pass)', 
                #status3='userid='+response_vat['userId']+ 'vrn='+response_vat['vrn'],
                status3='userid='+response_vat['userId']+ 'vrn='+response_vat['nino'], 
                status2=str(r.status_code) +' status reason'+ r.reason)
            l1.save()
            return render (request,  'pages/mtd/apis/tuagent_sa.html',
                    {'vat': 'create_testuser_sa'	,
                    'endpoint':  'user: ' + response_vat['userId'] + ' pwd:  ' + response_vat['password'],
                    'pd2':  str(r.status_code) + '  ' + r.reason,  #post_Data2,
                    'headers1': response_vat['nino'] , #headers,
                    'title': 'Testuser_sa   : PASS',
                    'clientid': obj.client_id,
                    'mtd_url': mtd_url })
        else: #non documented API status code
            a=1
            return render (request,  'pages/mtd/apis/tuagent_sa.html',
                    {'vat': 'create_testuser_sa'  ,
                    #'endpoint':  'user: ' + response_vat['userId'] + ' pwd:  ' + response_vat['password'],
                    #'pd2':  str(r.status_code) + '  ' + r.reason,  #post_Data2,
                    #'headers1': response_vat['nino  '] , #headers,
                    'title': 'Testuser_sa   : Undocumented Error Create TestUser SA',
                    'clientid': obj.client_id,
                    'mtd_url': mtd_url })
            #refACTOR THE CODE ABOVE!!!
#API to get last 30 days auth  which users have approved agents and gthis endpoint hasnt json
class a_last30(View):
    def get(self, request):
        tok=mtd_tokens.objects.latest('id')
        endpoint='https://test-api.service.hmrc.gov.uk/agents/'+tok.vrn +'/invitations'      

        post_Data2={
          }
        headers ={
            'Authorization': 'Bearer '+ tok.access_token,
            'Accept' :'application/vnd.hmrc.1.0+json',
            'Content-Type':'application/json',
            
        }

        r = requests.get(endpoint,data=post_Data2, headers=headers)
        qx={}
        x={}
        loc='None' #change header in template 
        #get invitation id from the response of this endpoint
        #put link on invitation id
        #link will then run next endpoint in   a_new_auth(View) below          
        l1=log1(src_sys_id='last30', status1=' LAST30', 
                status3='userid=' + endpoint,  
                status2=str(r.status_code) +' status reason '+ str(r.reason))
        l1.save()
        if r.status_code==204: 
            x= ' No clients  registered with this agent yet ' 
            l1.save()
        elif r.status_code==200:  #200 means we hv clients
            try:
                x=r.json()
            except (json.decoder.JSONDecodeError, KeyError, ValueError) as e:
                x='json  decode failed - app issue:' +str(e)
            else:
                loc='Yes'
                l1.save()
                i=1
                
                for q in x: #django templates cant use underscore in dict names 
                    qx[i]= q['_links']['self']['href']
                    i=i+1

        elif r.status_code==400 or r.status_code==401 or r.status_code==403 or r.status_code==406 or r.status_code==501 :        #elif  400 <= r.status_code  <= 499:
            try:   
                x= r.json()
            except (json.decoder.JSONDecodeError, KeyError, ValueError) as e:
                x='json  decode failed - hmrc sandbox unavailable with usual error message:' +str(e)
            else:
                l1=log1(src_sys_id='last30', status1=' LAST30', 
                status3='userid=' + endpoint,  
                status2=str(r.status_code) +' status reason '+ str(r.reason),
                status4=x['code'], status5=x['message'])
                l1.save()
        else: 
            x='Error in request with error code:' + str(r.status_code) + ' reason:  ' + str(r.reason)
        return render (request,  'pages/mtd/apis/tuagent.html',
                    {'vat': 'a_last30'    ,
                    'endpoint': endpoint,
                    'pd2':  str(r.status_code) + '  ' + str(r.reason),
                    'headers1': headers,
                    'title':  qx  ,  
                    #'clientid': obj.client_id, 
                    'location':loc,
                    'errors': x
                     })

class a_new_auth(View):
    #this is 2 part 
    #a)get - display form to enter clients vat details to be processed in (b) 
    #b)post - to take the form data in a) and register this client with hmrc and a link to be emailed ( in view  a_get_invite() below)
    template_name = 'pages/mtd/file_edit_auth_user.html'
    def get ( self, request):
        tok=mtd_tokens.objects.latest('id')
        pd={
              "service": ["MTD-VAT"],
              "clientType":"business",
            "clientIdType": "vrn",
            "clientId": "350970454",
            "knownFact": "2010-01-12",
            "arn": tok.vrn    #just pumps latest vrn in mtd_token table

          }
        return render (request,  self.template_name, pd)
    def post(self, request):
        if 'upload' in request.POST:
            tok=mtd_tokens.objects.latest('id')
            arn= request.POST.get('arn')
            endpoint='https://test-api.service.hmrc.gov.uk/agents/'+str(arn)+'/invitations'      

            post_Data2={
                  "service": ["MTD-VAT"] , #request.POST.get('service'),  cant get this to work as inout list from form,  so hardcode
                  "clientType":request.POST.get('clientType'),
                "clientIdType": request.POST.get('clientIdType'),
                "clientId": request.POST.get('clientId'),
                "knownFact": request.POST.get('knownFact'),

              }
            headers ={
                'Authorization': 'Bearer '+ tok.access_token,
                'Accept' :'application/vnd.hmrc.1.0+json',
                'Content-Type':'application/json',
                
            }

            r = requests.post(endpoint,data=json.dumps(post_Data2), headers=headers)
            if r.status_code==204:
                x='Location found',
                l1=log1(src_sys_id='a_new_auth', status1=tok.access_token, 
                    status3=request.POST.get('knownFact') ,  
                    status2=str(r.status_code) +' status reason '+ str( r.reason),
                    status4= str(request.POST.get('clientId')),
                    status5= 'arn: ' +str(request.POST.get('arn')),
                    status6='https://test-api.service.hmrc.gov.uk/'+r.headers['Location']
                    )
                l1.save()
            elif   r.status_code==400 or r.status_code==403 or r.status_code==406 or r.status_code==401 or r.status_code==500:    
                try:
                    y=r.json()
                except (json.decoder.JSONDecodeError, KeyError, ValueError) as e:
                    x='json  decode failed - hmrc sandbox down:' +str(e) 
                else:    
                    x='auth client failed with error code: ' + str( y['code']) + ' message: ' + y['message'] 
                    l1=log1(src_sys_id='a_new_auth', status1=tok.access_token, 
                    status3=request.POST.get('knownFact') ,  
                    status2=str(r.status_code) +' status reason '+ str( r.reason),
                    status4= str(request.POST.get('clientId')),
                    status5= 'arn: ' +str(request.POST.get('arn')),
                    status6='endpoint'
                    )
                    l1.save()
            else:
                l1=log1(src_sys_id='a_new_auth', status1=tok.access_token, 
                    status3=request.POST.get('knownFact') + x,  
                    status2=str(r.status_code) +' status reason '+ str( r.reason),
                    status4= str(request.POST.get('clientId')),
                    status5= 'arn: ' +str(request.POST.get('arn')),
                    status6=endpoint
                    )
                x= 'undocumented error in API'
                l1.save()
            return render (request,  'pages/mtd/apis/tuagent.html',
                        {'vat': 'new authorization'    ,
                        'endpoint': post_Data2,
                        'pd2':  str(r.status_code) + '  ' + str(r.reason),
                        'headers1': x,
                        'title': 'Post Authentication to get Link for Agent to send to client',
                        #'clientid': obj.client_id, 
                      #  'mtd_url': mtd_url
                         })


class a_get_invite(View):
    def get(self, request):
        tok=mtd_tokens.objects.latest('id')
        #endpoint='https://test-api.service.hmrc.gov.uk/'+client_link1    
        endpoint='https://test-api.service.hmrc.gov.uk/self-assessment/ni/'+tok.vrn+'/self-employments'
        post_Data2={
          "accountingPeriod": {
            "start": "2017-04-06",
            "end": "2018-04-05"
          },
          "accountingType": "CASH",
          "commencementDate": "2016-01-01",
          "tradingName": "Acme Ltd.",
          "businessAddressLineOne": "1 Acme Rd.",
          "businessAddressLineTwo": "London",
          "businessAddressLineThree": "Greater London",
          "businessAddressLineFour": "United Kingdom",
          "businessPostcode": "A9 9AA"
        }

        headers ={
            'Authorization': 'Bearer '+ tok.access_token,
            'Accept' :'application/vnd.hmrc.1.0+json',
            'Content-Type':'application/json',
            }

        r= requests.get(endpoint,data=post_Data2, headers=headers)
        y={}
        location=""
                   
        if r.status_code==201:
                try:
                    y=r.json()
                except (json.decoder.JSONDecodeError, KeyError, ValueError) as e:
                    x='json decode failed in SUCCESS -to investigate:' +str(e) 
                else:
                    
                        
                    x='Location found'
                    location='a_get_invite'
                    l1=log1(src_sys_id='a_new_auth', status1=tok.access_token, 
                       ## status3=request.POST.get('knownFact') ,  
                        status2=str(r.status_code) +' status reason '+ str( r.reason),
                       # status4= str(request.POST.get('clientId')),
                       # status5= 'arn: ' +str(request.POST.get('arn')),
                        status6=endpoint
                        )
                    l1.save()
                    if y['status'] =='Accepted':
                        return render (request,  'pages/mtd/apis/tuagent.html',
                            {'vat': 'new authorization_accepted'    ,
                            'endpoint': post_Data2,
                            'pd2':  str(r.status_code) + '  ' + str(r.reason),
                            'headers1': x,
                            'title': 'Post Authentication to get Link for Agent to send to client',
                            'location': location+'_accepted',
                            'y1':y,
                            'arn':  y['arn'],
                            'created': y['created'],
                            'updated':y['updated'],
                            'service':y['service'],
                            'status':y['status']
                            })
                    elif y['status'] =='Pending':
                        return render (request,  'pages/mtd/apis/tuagent.html',
                            {'vat': 'new authorization'    ,
                            'endpoint': post_Data2,
                            'pd2':  str(r.status_code) + '  ' + str(r.reason),
                            'headers1': x,
                            'title': 'Post Authentication to get Link for Agent to send to client',
                            'location': location,
                            'y1':y,
                            'links':  y['clientActionUrl'],
                            'created': y['created'],
                            'expiresOn':y['expiresOn'],
                            'service':y['service'],
                            'status':y['status']
                            #'clientid': obj.client_id, 
                          #  'mtd_url': mtd_url
                             })
        elif   r.status_code==400 or r.status_code==403 or r.status_code==404 or r.status_code==406 or r.status_code==500 or r.status_code==401:    
                try:
                    y=r.json()
                except (json.decoder.JSONDecodeError, KeyError, ValueError) as e:
                    x='json  decode failed in  ERROR-to investigate:' +str(e) 
                else:    
                    x='auth client failed with error code: ' + str( y['code']) + ' message: ' + y['message'] 
                    l1=log1(src_sys_id='get_invite_by_id_endpoint', status1=tok.access_token, 
                    #status3=request.POST.get('knownFact') ,  
                    status2=str(r.status_code) +' status reason '+ str( r.reason),
                    #status4= str(request.POST.get('clientId')),
                    #status5= 'arn: ' +str(request.POST.get('arn')), #should add arn of agent -so know who agent is??
                    status6='endpoint'
                    )
                    l1.save()
                    x='expected error in endpoint'
        else:   
                l1=log1(src_sys_id='a_new_auth', status1=tok.access_token, 
                    #status3=request.POST.get('knownFact') + x,  
                    status2=str(r.status_code) +' status reason '+ str( r.reason),
                    #status4= str(request.POST.get('clientId')),
                    #status5= 'arn: ' +str(request.POST.get('arn')),
                    status6=endpoint
                    )
                x= 'undocumented error in API'
                l1.save()
        return render (request,  'pages/mtd/apis/tuagent.html',
                        {'vat': 'new authorization'    ,
                        'endpoint': post_Data2,
                        'pd2':  str(r.status_code) + '  ' + str(r.reason),
                        'headers1': x,
                        'title': 'Post Authentication to get Link for Agent to send to client',
                        'location': location,
                        'y1':y
                        
                         })

class a_get_relationship(View):
    #this is 2 part 
    #a)get - display form to enter clients vat details to be processed in (b) 
    #b)post - to take the form data in a) and register this client with hmrc and a link to be emailed ( in view  a_get_invite() below)
    template_name = 'pages/mtd/file_edit_auth_relship.html'
    def get ( self, request):
        tok=mtd_tokens.objects.latest('id')
        pd={
              "service": ["MTD-VAT"],
              "clientType":"business",
            "clientIdType": "vrn",
            "clientId": "350970454",
            "knownFact": "2010-01-12",
            "arn": tok.vrn    #just pumps latest vrn in mtd_token table

          }
        return render (request,  self.template_name, pd)
    def post(self, request):
        if 'upload' in request.POST:
            tok=mtd_tokens.objects.latest('id')
            arn= request.POST.get('arn')
            endpoint='https://test-api.service.hmrc.gov.uk/agents/'+str(arn)+'/relationships'      

            post_Data2={
                  "service": ["MTD-VAT"] , #request.POST.get('service'),  cant get this to work as inout list from form,  so hardcode
           #       "clientType":request.POST.get('clientType'),
                "clientIdType": request.POST.get('clientIdType'),
                "clientId": request.POST.get('clientId'),
                "knownFact": request.POST.get('knownFact')

              }
            headers ={
                'Authorization': 'Bearer '+ tok.access_token,
                'Accept' :'application/vnd.hmrc.1.0+json',
                'Content-Type':'application/json',
                
            }

            r = requests.post(endpoint,data=json.dumps(post_Data2), headers=headers)
            if r.status_code==204:
                x='204: There exist client -agent relationship',
                l1=log1(src_sys_id='a_new_rel', status1=tok.access_token, 
                    status3=request.POST.get('knownFact') ,  
                    status2=str(r.status_code) +' status reason '+ str( r.reason),
                    status4= str(request.POST.get('clientId')),
                    status5= 'arn: ' +str(request.POST.get('arn')),
                    status6='https://test-api.service.hmrc.gov.uk/'#+r.headers['Location']
                    )
                l1.save()
                #location ='YEs Exists Agent - Client Relationship'
            elif   r.status_code==400 or r.status_code==403 or r.status_code==406 or r.status_code==401 or r.status_code==500 or r.status_code==404:    
                try:
                    y=r.json()
                except (json.decoder.JSONDecodeError, KeyError, ValueError) as e:
                    x='json  decode failed - internal svr issue:' +str(e) 
                 #   location
                else:    
                    x='auth relationship failed with error code: ' + str( y['code']) + ' message: ' + y['message'] 
                    l1=log1(src_sys_id='a_new_rel', status1=tok.access_token, 
                    status3=request.POST.get('knownFact') ,  
                    status2=str(r.status_code) +' status reason '+ str( r.reason),
                    status4= str(request.POST.get('clientId')),
                    status5= 'arn: ' +str(request.POST.get('arn')),
                    status6='endpoint'
                    )
                    l1.save()
            else:
                x= ' error without status code in API' 
                l1=log1(src_sys_id='a_new_rel', status1=tok.access_token, 
                    status3=request.POST.get('knownFact') + x,  
                    status2=str(r.status_code) +' status reason '+ str( r.reason),
                    status4= str(request.POST.get('clientId')),
                    status5= 'arn: ' +str(request.POST.get('arn')),
                    status6=endpoint
                    )
                
                l1.save()
            return render (request,  'pages/mtd/apis/tuagent.html',
                        {'vat': 'new agent relationship'    ,
                        'endpoint': post_Data2,
                        'pd2':  str(r.status_code) + '  ' + str(r.reason),
                        'headers1': x,
                        'title': 'Get Agent relationship for  client',
                        'location': 'a_get_relationship'
                        #'clientid': obj.client_id, 
                      #  'mtd_url': mtd_url
                         })


class a_cancel_invite(View):
    def get(self, request):
        tok=mtd_tokens.objects.latest('id')
        endpoint='https://test-api.service.hmrc.gov.uk/agents/CARN0466671/invitations'      

        post_Data2={
          }
        headers ={
            'Authorization': 'Bearer '+ tok.access_token,
            'Accept' :'application/vnd.hmrc.1.0+json',
            'Content-Type':'application/json',
            
        }

        response_vat1 = requests.get(endpoint,data=post_Data2, headers=headers)

                   
        l1=log1(src_sys_id='last30', status1=' LAST30', 
                status3='userid=' + endpoint,  
                status2=str(response_vat1.status_code) +' status reason '+ str( response_vat1.reason))
        l1.save()
        return render (request,  'pages/mtd/apis/tuagent.html',
                    {'vat': 'a_cancel_invite'    ,
                    'endpoint': endpoint,
                    'pd2':  str(response_vat1.status_code) + '  ' + str(response_vat1.reason),
                    'headers1': headers,
                    'title': 'suitable messahe for this view'
                    #'clientid': obj.client_id, 
                  #  'mtd_url': mtd_url
                     })



class refresh_token(View):
    def get(self, request):
        tok=mtd_tokens.objects.latest('id')
        endpoint = "https://test-api.service.hmrc.gov.uk/oauth/token"

        pd ={
            'client_secret' :'a97c7346-a779-403c-ba18-34abb71374c8',
            'client_id':'fQ84pQWgR71bsRWaHRSKRFsh5JQa',
            'grant_type':'refresh_token',
            'refresh_token':tok.refresh_token
        }
        r = requests.post(endpoint,data=pd)
        l1=log1(src_sys_id='refresh token', status1=' refresh token', 
                status3='endpoint=' + endpoint,  
                status2=str(r.status_code) +' status reason '+ str( r.reason))
        l1.save()
        if r.status_code == 200:
            xx= 'refresh ok'
            content=r.json()
            t=mtd_tokens(
                access_token=content['access_token'],
                access_code='ac1',
                refresh_token=content['refresh_token'], 
                scope=content['scope'],     
                endpoint='https://test-api.service.hmrc.gov.uk/oauth/token',
                vrn=tok.vrn,   
                src_sys_id='refresh_token',
                userid= 'na',
                password='na'   )
            t.save()
        #refresh token api provides detailed 4xx errors in json format with keys error and error_description
        #other api end points provide across different keys and some not in json
        #why so inconsistent
        elif  r.status_code  == 400:
            x= r.json()
            xx="error: " + x["error"] + " error description: " + x["error_description"]
        else: 
            xx='refresh token failed with status_code' + str( r.status_code)
        return render (request,  'pages/mtd/apis/tuagent.html',
                    {'vat': 'refresh access_token using refresh_token for arn/vrn: ' + tok.vrn    ,
                    'endpoint': endpoint,
                    'pd2':  str(r.status_code) + '  ' + str(r.reason),
                    'headers1': 'no headers',
                    'title': xx,
                    #'clientid': obj.client_id, 
                  #  'mtd_url': mtd_url
                     })



class MtdView2(View):
    def get(self, request):
        obj=apps_ids.objects.get(id=10)
        code1=request.GET.get('code','')
        vrn1=request.GET.get('state','')
        parsed_uri =     urlparse(request.build_absolute_uri())
        base_url = '{uri.scheme}://{uri.netloc}/'.format(uri=parsed_uri)
        if 'https' in base_url:
            mtd_url = base_url
        elif 'http' in base_url:
            mtd_url = base_url.replace("http","https")
        else:
            mtd_url = base_url
        post_data={'client_secret':obj.client_Secrets,
        'client_id': obj.client_id ,'grant_type':'authorization_code',
        'redirect_uri':mtd_url+'mtd/mtd2agent/',
        # 'redirect_uri':'https://hmrcfinal.herokuapp.com/mtd/mtd2/',
        # 'redirect_uri':'http://localhost:8000/mtd/mtd2/',
        #'state': vrn1,
        #takeout as not needed in api plus this hold vrn
        'code': code1 }


        response = requests.post('https://test-api.service.hmrc.gov.uk/oauth/token', data=post_data)


        #post auth code to get access token
        try:
            content= response.json()
        except (json.decoder.JSONDecodeError, KeyError, ValueError) as e: 
            l1=log1(src_sys_id='Testuser', status1=' Try  JSON decode error' + str(e),
                status2=str(response_vat1.status_code) +' status reason'+ response_vat1.reason)
            l1.save()
            return render (request,  'pages/mtd/apis/tu.html',
                    {#'vat': response_vat   ,
                    'endpoint': endpoint,
                    'pd2':post_Data2,
                    'headers1': headers,
                    'title': 'TestUser: Failed MTD token- check logs',
                    'clientid': obj.client_id,
                    'mtd_url': mtd_url })   
        
        else:    
            t=mtd_tokens(
            access_token=content['access_token'],
            access_code='ac1',
            refresh_token=content['refresh_token'],	
            scope=content['scope'],		
            endpoint='https://test-api.service.hmrc.gov.uk/oauth/token',
            vrn=vrn1,#this come manually from tuagent template in the link @state
            #vrn=content['agentServicesAccountNumber'],  #update here for agent vrn  #standardize 'content' to 'r'
            src_sys_id='mtd_views2_agent',
            userid= 'na',
            password='na'	)
            t.save()

            return render(request, 'pages/mtd/apis/mtd_home2.html',{ 'code': code1,
            'content': content , 
            'vat': 'updated list_mtd db',
            'endpoint': 'oath/token',
            'pd2':'post_Data3',
            'headers1':code1,	
            'comment': 'Agent arn ' + vrn1,  #update here to display avrn
                }

        )    

class VatLiabilitiesView(View):
    def get(self, request):
        tok=mtd_tokens.objects.latest('id')
        obj=apps_ids.objects.get(id=5)
        posted_obj = posted.objects.latest('id') #filter(vrn=tok.vrn)[:0].get()
        periodpj=posted_obj.periodKey
        tokpj=tok.access_token



        # generating access token code
        endpoint1 = "https://test-api.service.hmrc.gov.uk/oauth/token"
        print(tok.refresh_token)
        post_data ={
            'client_secret' :'f18af97f-2acb-41f9-92d2-e373c44618a8',
            'client_id':'ZjScpZHoNA6TQJytfrEkFcSryf4a',
            'grant_type':'refresh_token',
            'refresh_token':tok.refresh_token
        }

        response_vat1 = requests.post(endpoint1, data=post_data)
        print('oath api')
        print(response_vat1)
        oath_api_response = response_vat1.json()
        print(oath_api_response)
        if 'error' in oath_api_response: 
            api_status = {'status_code': oath_api_response['error'], 'reason':oath_api_response['error_description']}
            print('coming in auth error')
            return render(request, 'pages/mtd/subnav/vat_liabilities.html', { 'result': oath_api_response , 'full_path': request.get_full_path(),'vrn':tok.vrn, 'api_status': api_status})
        else:
            print('auth successful')
            endpoint='https://test-api.service.hmrc.gov.uk/organisations/vat/'+tok.vrn +'/liabilities' 

            #endpoint='https://test-api.service.hmrc.gov.uk/organisations/vat/'+tok.vrn +'/obligations/'

            headers ={
                'Authorization': 'Bearer '+ oath_api_response["access_token"],  #'5c5adcd9d4ee25ac173c78225ef45462',
                'Accept' :'application/vnd.hmrc.1.0+json',
                #'Gov-Test-Scenario':'QUARTERLY_NONE_MET',
                #	'Gov-Test-Scenario':'INVALID_PERIODKEY',

                'Content-Type':'application/json'#,
                #"from": "2018-01-01", 
                # "to": "2018-12-31"
            }


            post_Data2={
            'from': '2019-06-14', 
            'to': '2020-05-29',
            #'status':'F'	  
                    }


            l= requests.get(endpoint,data=json.dumps(post_Data2), headers=headers)
            if l.ok:

                print(l.request.body)
                print(l.request.headers)
                print(l.request.url)
                try:
                    lj=l.json()
                except:
                    print('failed json in try/except')
                else:
                    print('PASS json details ' + lj)
                    api_status = {'status_code': l.status_code, 'reason':l.reason}
                    return render(request, 'pages/mtd/subnav/vat_liabilities.html', {'result': lj, 'full_path': request.get_full_path(), 'vrn':tok.vrn, 'api_status': api_status})

            else:
                error = l.json()
                print('failed get l.ok is:' + str(l.ok) + ' status code:' + str(l.status_code) + ' reason code: ' + str(l.reason))
                api_status = {'status_code': error['statusCode'], 'reason':error['message']}
                print ('detailed fail reason: ')
                print( l.json())        
                return render(request, 'pages/mtd/subnav/vat_liabilities.html', { 'result': l , 'full_path': request.get_full_path(),'vrn':tok.vrn, 'api_status': api_status})

class VatPaymentView(View):
    def get(self, request):
        tok=mtd_tokens.objects.latest('id')
        obj=apps_ids.objects.get(id=5)
        posted_obj = posted.objects.latest('id') #filter(vrn=tok.vrn)[:0].get()
        periodpj=posted_obj.periodKey
        tokpj=tok.access_token


        endpoint='https://test-api.service.hmrc.gov.uk/organisations/vat/'+tok.vrn +'/payments'#+ periodpj #"%23001"  

        #endpoint='https://test-api.service.hmrc.gov.uk/organisations/vat/'+tok.vrn +'/obligations/'

        headers ={
            'Authorization': 'Bearer '+ tok.access_token,  #'5c5adcd9d4ee25ac173c78225ef45462',
            'Accept' :'application/vnd.hmrc.1.0+json',
            #'Gov-Test-Scenario':'QUARTERLY_NONE_MET',
            #	'Gov-Test-Scenario':'INVALID_PERIODKEY',

            'Content-Type':'application/json'
        }

        post_Data2={
        #'to': '2019-06-15',

        #'from': '2019-06-12' 
        #'status':'F'	  
                }

        p=[]
        pj=[]		
        p= requests.get(endpoint,data=json.dumps(post_Data2), headers=headers)
                




        print(p.request.body)


        print(p.request.headers)

        print(p.request.url)

        pj=p.json()
        return render(request, 'pages/mtd/subnav/vat_payment.html', {'result': pj, 'full_path': request.get_full_path()})


class VatObligationsView(View):
    def get(self, request):
        tok=mtd_tokens.objects.latest('id')
        obj=apps_ids.objects.get(id=5)
        #posted_obj = posted.objects.filter(vrn=tok.vrn)[:1].get()
        #periodvr=posted_obj.periodKey
        tokvj=tok.access_token

        #endpoint='https://test-api.service.hmrc.gov.uk/organisations/vat/'+tok.vrn +'/returns/'+ period  

        endpoint='https://test-api.service.hmrc.gov.uk/organisations/vat/'+tok.vrn +'/obligations'

        headers ={
            'Authorization': 'Bearer '+ tok.access_token,  #'5c5adcd9d4ee25ac173c78225ef45462',
            'Accept' :'application/vnd.hmrc.1.0+json',
            #'Gov-Test-Scenario':'INVALID_PERIODKEY',
            'Content-Type':'application/json'
        }

        #date optional to removing
        post_Data2={
        "from": "2018-06-14", 
        "to": "2018-06-15",
        "status":"F"	  
                }
        l=[]
                #json.dumps(post_Data2)

        o= requests.get(endpoint,data=json.dumps(post_Data2), headers=headers)
                


        print ("print(o.request.body)")
        print(o.request.body)

        print(o.request.headers)

        print(o.request.url)
        print(o.json())

        try:
            print('coming in try')
            api_status = {'status_code': o.status_code, 'reason':o.reason}
            oj=o.json()
        except:
            print('coming in except')
            api_status = {'status_code': o.status_code, 'reason':o.reason}

        return render(request, 'pages/mtd/subnav/vat_obligations.html', {'result': oj, 'full_path': request.get_full_path(), 'vrn':tok.vrn, 'api_status': api_status})

class VatReturnView(View):
    def get(self, request):
        tok=mtd_tokens.objects.latest('id')
        obj=apps_ids.objects.get(id=5)
        posted_obj = posted.objects.latest('id') #filter(vrn=tok.vrn)[:0].get()
        periodpj=posted_obj.periodKey
        tokpj=tok.access_token


        endpoint='https://test-api.service.hmrc.gov.uk/organisations/vat/'+tok.vrn +'/returns/'+ periodpj #"%23001"  

        #endpoint='https://test-api.service.hmrc.gov.uk/organisations/vat/'+tok.vrn +'/obligations/'

        headers ={
            'Authorization': 'Bearer '+ tok.access_token,  #'5c5adcd9d4ee25ac173c78225ef45462',
            'Accept' :'application/vnd.hmrc.1.0+json',
            #'Gov-Test-Scenario':'QUARTERLY_NONE_MET',
            #	'Gov-Test-Scenario':'INVALID_PERIODKEY',

            'Content-Type':'application/json'
        }


        post_Data2={
        'from': '2012006-30', 
        'to': '2020-05-31',
        #'status':'F'	  
                }

        r= requests.get(endpoint,data=json.dumps(post_Data2), headers=headers)
                


        print(r.request.body)


        print(r.request.headers)

        print(r.request.url)
        try:
            print('coming in try')
            print(str(r.status_code) +r.reason)
            api_status = {'status_code': r.status_code, 'reason':r.reason}
            rj=r.json()
        except:
            print('coming in except')
            print(str(r.status_code) +r.reason)
            api_status = {'status_code': r.status_code, 'reason':r.reason}

        return render(request, 'pages/mtd/subnav/vat_return.html', {'result': rj, 'full_path': request.get_full_path(), 'vrn':tok.vrn, 'api_status': api_status}) 