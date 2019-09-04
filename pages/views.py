from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.views import View

from .s3_handler import S3Handler
from openpyxl import load_workbook, Workbook
import requests
import os
import time
import json
import random
from .models import mtd_tokens,posted ,testuser, apps_ids, log1, auth_users, invoice, products, payment_status
from urllib.parse import urlparse
# Create your views here.
class HomeView(View):
    def get(self, request):
        print('home')
        # return HttpResponse('homeee')
        
        return render(request, 'pages/index.html', {'full_path': request.get_full_path()})

class displayall(View):
    template_name='pages/mtd/apis/displayall.html'
    def get(self,request):
        #apps=apps_ids.objects.all()
        #testuser1=testuser.objects.all().order_by('-id')[:5]
        #mtd_tokens1=mtd_tokens.objects.all().order_by('-id')[:5]
        #posted1=posted.objects.all().order_by('-id')[:10]
        log1all=log1.objects.all().order_by('-id')[:10]
        au=auth_users.objects.all().order_by('-id')
        
        return render (request,self.template_name,
                    {#'vat': apps,   
                    #'testuser': testuser1,
                    #'mtd_tokens': mtd_tokens1,
                    #'posted': posted1,
                    'log1': log1all,
                    'au': au,
                    'title': 'Displayall: Refresh  access token  Return'})  

class MtdView(View):
    template_name = 'pages/mtd/uploadfile.html'

    def get(self, request):
        s3_handler = S3Handler()
        file_list = s3_handler.get_file_list()
        return render(request, self.template_name,{'vat': file_list, 'full_path': request.get_full_path()})

class AgentView(View):
    template_name = 'pages/mtd/agentview.html'

    def get(self, request):
       # s3_handler = S3Handler()
       # file_list = s3_handler.get_file_list()
        return render(request, self.template_name )#,{'vat': file_list, 'full_path': request.get_full_path()})

class ClientView(View):
    template_name = 'pages/mtd/agentview_clientview.html'

    def get(self, request):
       # s3_handler = S3Handler()
       # file_list = s3_handler.get_file_list()
        return render(request, self.template_name )#,{'vat': file_list, 'full_path': request.get_full_path()})


class AgentView_sa(View):
    template_name = 'pages/mtd/agentview_sa.html'

    def get(self, request):
       # s3_handler = S3Handler()
       # file_list = s3_handler.get_file_list()
        return render(request, self.template_name )#,{'vat': file_list, 'full_path': request.get_full_path()})

class FileListView(View):
    
    def get(self, request):

       # if not request.user.is_authenticated:
        #    return JsonResponse(status=401)

        s3_handler = S3Handler()
        file_list = s3_handler.get_file_list()

        return JsonResponse(data={'data': file_list}, safe=False)

class UploadFileView(View):

    def post(self, request):
        file = request.FILES['file']  
        file_name = file.name.split(".")[0]+'_'+str(random.randint(9999,9999999))
        full_file_name = file_name+os.path.splitext(file.name)[1]
        
        # wb = load_workbook(filename=file)

        # file_path = "%s%s" % (time.time(), os.path.splitext(file.name)[1])

        # wb.save(file_path)

        wb = load_workbook(filename=file)
        file_path = "%s%s" % (time.time(), os.path.splitext(full_file_name)[1])
        wb.save(file_path)
        s3_handler = S3Handler()
        s3_handler.upload_file(file_path, full_file_name)
        os.remove(file_path)
        #updated to it goes to 'mtd' path which is the intended
        return redirect('/mtd/')


class FileEditView(View):
    template_name = 'pages/mtd/file_edit.html'
    
    def get(self, request):
            
        api_response_code = None
        api_response_message = None
        if request.GET.get('api_response_code'):
            api_response_code = request.GET.get('api_response_code')
        if request.GET.get('api_response_message'):
            api_response_message = request.GET.get('api_response_message')
        key = request.GET.get('key')
        s3_handler = S3Handler()
        download_path = s3_handler.download_file(key)
        wb = load_workbook(download_path)
        ws = wb[wb.sheetnames[0]]
        data = []
        for col in ws[1]:
            data.append({'header': col.value})

        for col in ws[2]:
            data[col.column - 1]['value'] = col.value
        os.remove(download_path)
        return render(request, self.template_name, {'data': data, 'key': key, 'api_response_code': api_response_code, 'api_response_message': api_response_message})

    

    def post(self, request):
        if 'save' in request.POST:

            key = request.POST.get('key')
            headers = ['periodKey', 'vatDueSales', 'vatDueAcquisitions', 'totalVatDue', 'vatReclaimedCurrPeriod',
                    'netVatDue', 'totalValueSalesExVAT', 'totalValuePurchasesExVAT', 'totalValueGoodsSuppliedExVAT',
                    'totalAcquisitionsExVAT', 'finalised']
            wb = Workbook()
            ws = wb.active
            ws.append(headers)
            data = []
            for header in headers:
                data.append(request.POST.get(header))
            ws.append(data)
            
            file_path = "%s%s" % (time.time(), os.path.splitext(key)[1])
            wb.save(file_path)
            
            s3_handler = S3Handler()
            s3_handler.upload_file(file_path, key)
            os.remove(file_path)
            return redirect('/mtd/')
        else:
            periodKey = request.POST.get('periodKey')
            vatDueSales = request.POST.get('vatDueSales')
            vatDueAcquisitions = request.POST.get('vatDueAcquisitions')
            totalVatDue = request.POST.get('totalVatDue')
            vatReclaimedCurrPeriod = request.POST.get('vatReclaimedCurrPeriod')
            netVatDue = request.POST.get('netVatDue')
            totalValueSalesExVAT = request.POST.get('totalValueSalesExVAT')
            totalValuePurchasesExVAT = request.POST.get('totalValuePurchasesExVAT')
            totalValueGoodsSuppliedExVAT = request.POST.get('totalValueGoodsSuppliedExVAT')
            totalAcquisitionsExVAT = request.POST.get('totalAcquisitionsExVAT')
            finalised = request.POST.get('finalised')
            
            #tok=mtd_tokens.objects.get(id=8)
            tok=mtd_tokens.objects.latest('id')
            print('tok')
            print(tok)
            print(periodKey)
            endpoint='https://test-api.service.hmrc.gov.uk/organisations/vat/'+tok.vrn+'/returns'
        

            #vrn='700695645'
            #endpoint='https://test-api.service.hmrc.gov.uk/organisations/vat/'+vrn+'/returns'
            post_Data3={
                "periodKey": periodKey,
                "vatDueSales": vatDueSales,
                "vatDueAcquisitions": vatDueAcquisitions,
                "totalVatDue": totalVatDue,
                "vatReclaimedCurrPeriod": vatReclaimedCurrPeriod,
                "netVatDue": netVatDue,
                "totalValueSalesExVAT": totalValueSalesExVAT,
                "totalValuePurchasesExVAT": totalValuePurchasesExVAT,
                "totalValueGoodsSuppliedExVAT": totalValueGoodsSuppliedExVAT,
                "totalAcquisitionsExVAT": totalAcquisitionsExVAT,
                "finalised": True
                }
            headers ={
                    #"Authorization": "Bearer 61d270d116111693fd3fa8e061b9753c",
                    "Authorization": "Bearer " + tok.access_token,
                    "Accept" :"application/vnd.hmrc.1.0+json",
                    "Content-Type":"application/json"
                #"Gov-Client-Connection-Method": "DESKTOP_APP_VIA_SERVER"
            }
            response_vat1 = requests.post(endpoint,data=json.dumps(post_Data3), headers=headers)
            key = request.POST.get('key')

            try:
                response_vat=response_vat1.json()
            #except:
            except (json.decoder.JSONDecodeError, KeyError, ValueError) as e: 

                response_vat ={'code': 'DEAD', 'message': 'Server Issue'}
                print ('response json error in PostVat')
                l1=log1(src_sys_id='PostVat2', status1=' fail: Json deode Error '+ str(e) , 
                status2= str(response_vat1.status_code) +response_vat1.reason)
                l1.save()
                
            else:
                if 'code' in response_vat:
                    #zz=response_vat["code"]
                    zz=response_vat
                    print('zz')
                    print(zz)
                    # print(zz['code'])
                    # yy=zz["errors"]
                    
                    l1=log1(src_sys_id='PostVat2', status1=' fail JSON ok but error +vrn=' + str(tok.vrn), 
                        status2=response_vat["code"] +' '+ zz['code']+'  '+ zz['message']  , 
                        status3=str(response_vat1.status_code) +response_vat1.reason +'post period' +post_Data3['periodKey'] )
                    l1.save()
                    return render (request, 'pages/mtd/mtd_home2.html',{ 
                         'vat': response_vat,
                         'comment': 'PostVat JSON error in response and code is:' + tok.vrn
                         })
                else: #assume we are posting as no json decode error or post json decode 'error'
                        #if response_vat1
                        t1=posted( periodKey=post_Data3['periodKey'],               vrn=tok.vrn,
                        vatDueSales=post_Data3['vatDueSales'],           
                           vatDueAcquisitions=post_Data3["vatDueAcquisitions"],
                        #totalVatDue=post_Data3["totalVatDue"],
                        vatReclaimedCurrPeriod=post_Data3["vatReclaimedCurrPeriod"],    
                                    netVatDue=post_Data3["netVatDue"],
                        totalValueSalesExVAT=["totalValueSalesExVAT"],              totalValuePurchasesExVAT=post_Data3["totalValuePurchasesExVAT"],
                        totalValueGoodsSuppliedExVAT=["totalValueGoodsSuppliedExVAT"],
                                      totalAcquisitionsExVAT=["totalAcquisitionsExVAT"],
                        finalised=post_Data3["finalised"],
                                      access_token=response_vat["processingDate"],
                        endpoint=response_vat["paymentIndicator"],
                                      date=response_vat["formBundleNumber"]
                        #vrn=response_vat["chargeRefNumber"]

                        )
                        t1.save()
                    
                        l1=log1(src_sys_id='PostVat2', status1=' else:success ',
                            status3=str(response_vat1.status_code) +response_vat1.reason,
                            status2=tok.vrn)
                        l1.save()
                        #lj take this out - no 'code' as this branch ot try/except/else has no error
                       #return redirect('file-edit?key='+key+'&api_response_code='+response_vat["code"]+'&api_response_message='+response_vat["message"]+'')
                        return redirect('/mtd/')

class TestView(View):
    def get(self, request):
        # tok=mtd_tokens.objects.latest('id')
        # tok = '99a47b444c236f4e673c19d3961f1d'
        # print('tok')
        # print(tok)
        # obj=apps_ids.objects.get(id=5)
        # posted_obj = posted.objects.latest('id') #filter(vrn=tok.vrn)[:0].get()
        # periodpj=posted_obj.periodKey
        periodpj="0001"
        tokpj='9b51aaa6f3cc176bff8917e8145d528'
    

        endpoint='https://test-api.service.hmrc.gov.uk/organisations/vat/'+'997382159' +'/returns/'+ periodpj #"%23001"  

        #endpoint='https://test-api.service.hmrc.gov.uk/organisations/vat/'+tok.vrn +'/obligations/'

        headers ={
        'Authorization': 'Bearer '+ tokpj,  #'5c5adcd9d4ee25ac173c78225ef45462',
        'Accept' :'application/vnd.hmrc.1.0+json',
        #'Gov-Test-Scenario':'QUARTERLY_NONE_MET',
        #	'Gov-Test-Scenario':'INVALID_PERIODKEY',

        'Content-Type':'application/json'
        }


        post_Data2={
        'from': '2019-06-30', 
        'to': '2020-05-31',
        #'status':'F'	 
        }

        r= requests.get(endpoint,data=json.dumps(post_Data2), headers=headers)



        print(r.request.body)


        print(r.request.headers)

        print(r.request.url)
        print(r)
        print('abc')
        rj=r.json()

        print(rj)
        return HttpResponse('abc')


class InitView(View):
    def get(self, request):
        t= apps_ids(appname= 'bts1',		
                    client_Secrets='f18af97f-2acb-41f9-92d2-e373c44618a8',	
                    client_id='ZjScpZHoNA6TQJytfrEkFcSryf4a',
                    server_token='9d41fd904fc3e6449a51381f76e43df'
                    
                )
        t.save()
        t= apps_ids(appname= 'bts1test',        
                    client_Secrets='a97c7346-a779-403c-ba18-34abb71374c8',  
                    client_id='fQ84pQWgR71bsRWaHRSKRFsh5JQa',
                    server_token='98918245c5488ac09cf5a55955c5e1cd'
                    
                )
        t.save()
        parsed_uri = urlparse(request.build_absolute_uri())
        base_url = '{uri.scheme}://{uri.netloc}/'.format(uri=parsed_uri)+'mtd/mtd2/'
        
        if 'https' in base_url:
            mtd_url = base_url
        elif 'http' in base_url:
            mtd_url = base_url.replace("http","https")
        else:
            mtd_url = base_url
        
        return render(request, 'pages/mtd/apis/tu.html',{'mtd_url':mtd_url})

class TuView(View):
    def get(self, request):
        obj=apps_ids.objects.get(id=10)
        parsed_uri = urlparse(request.build_absolute_uri())
        base_url = '{uri.scheme}://{uri.netloc}/'.format(uri=parsed_uri)+'mtd/mtd2/'
        
        if 'https' in base_url:
            mtd_url = base_url
        elif 'http' in base_url:
            mtd_url = base_url.replace("http","https")
        else:
            mtd_url = base_url
        endpoint = 'https://test-api.service.hmrc.gov.uk/create-test-user/organisations'		
        #endpoint='https://test-api.service.hmrc.gov.uk/create-test-user/agents'		

        post_Data2={
            "serviceNames":   #  ["agent-services"] 
             ["mtd-vat"]
        }
        headers ={
            'Authorization': 'Bearer '+ obj.server_token,
            'Accept' :'application/vnd.hmrc.1.0+json',
            'Content-Type':'application/json',
            
        }

        response_vat1 = requests.post(endpoint,data=json.dumps(post_Data2), headers=headers)

        try:
            response_vat=response_vat1.json()
        except (json.decoder.JSONDecodeError, KeyError, ValueError) as e: 
            l1=log1(src_sys_id='Testuser', status1=' Try  JSON decode error' + str(e),
                status2=str(response_vat1.status_code) +' status reason'+ response_vat1.reason)
            l1.save()
            return render (request,  'pages/mtd/apis/tu.html',
                    {#'vat': response_vat	,
                    'endpoint': endpoint,
                    'pd2':post_Data2,
                    'headers1': headers,
                    'title': 'TestUser: Failed JD - check logs'+ str(response_vat1.status_code)
                        + str( response_vat1.reason),
                    'clientid': obj.client_id,
                    'mtd_url': mtd_url })	
        else:
            t2=testuser(endpoint=endpoint,
                vrn=response_vat['vrn'],
                #vrn=response_vat['agentServicesAccountNumber'],
                userid= response_vat['userId'],
                password=response_vat['password'],
                src_sys_id='tu.view.tu',
                appname=obj.appname 
                )
            t2.save()
            l1=log1(src_sys_id='Testuser', status1=' Try  else(pass)', 
                status3='userid='+response_vat['userId']+ 'vrn='+response_vat['vrn'],
                #status3='userid='+response_vat['userId']+ 'vrn='+response_vat['agentServicesAccountNumber'], 
                status2=str(response_vat1.status_code) +' status reason'+ response_vat1.reason)
            l1.save()
            return render (request,  'pages/mtd/apis/tu.html',
                    {'vat': response_vat	,
                    'endpoint': endpoint,
                    'pd2':  str(response_vat1.status_code) + '  ' + response_vat1.reason,  #post_Data2,
                    'headers1': headers,
                    'title': 'TestUser: data with db',
                    'clientid': obj.client_id,
                    'mtd_url': mtd_url })

class MtdView2(View):
    def get(self, request):
        obj=apps_ids.objects.get(id=10)
        code1=request.GET.get('code','')
        vrn1=request.GET.get('state','')
        parsed_uri = urlparse(request.build_absolute_uri())
        base_url = '{uri.scheme}://{uri.netloc}/'.format(uri=parsed_uri)
        if 'https' in base_url:
            mtd_url = base_url
        elif 'http' in base_url:
            mtd_url = base_url.replace("http","https")
        else:
            mtd_url = base_url
        post_data={'client_secret':obj.client_Secrets,
        'client_id': obj.client_id ,'grant_type':'authorization_code',
        'redirect_uri':mtd_url+'mtd/mtd2/',
        # 'redirect_uri':'https://hmrcfinal.herokuapp.com/mtd/mtd2/',
        # 'redirect_uri':'http://localhost:8000/mtd/mtd2/',
        #'state': vrn1,
        #takeout as not needed in api plus this hold vrn
        'code': code1 }


        response = requests.post('https://test-api.service.hmrc.gov.uk/oauth/token', data=post_data)


        #post auth code to get access token
        try:
            content= response.json()
        except (json.decoder.JSONDecodeError, KeyError, ValueError) as e: 
            l1=log1(src_sys_id='Testuser', status1=' Try  JSON decode error' + str(e),
                status2=str(response_vat1.status_code) +' status reason'+ response_vat1.reason)
            l1.save()
            return render (request,  'pages/mtd/apis/tu.html',
                    {#'vat': response_vat   ,
                    'endpoint': endpoint,
                    'pd2':post_Data2,
                    'headers1': headers,
                    'title': 'TestUser: Failed MTD token- check logs',
                    'clientid': obj.client_id,
                    'mtd_url': mtd_url })   
        
        else:    
            t=mtd_tokens(
            access_token=content['access_token'],
            access_code='ac1',
            refresh_token=content['refresh_token'],	
            scope=content['scope'],		
            endpoint='https://test-api.service.hmrc.gov.uk/oauth/token',
            vrn=vrn1,	
            src_sys_id='mtd_views2',
            userid= 'na',
            password='na'	)
            t.save()

            return render(request, 'pages/mtd/apis/mtd_home2.html',{ 'code': code1,
            'content': content , 
            'vat': 'updated list_mtd db',
            'endpoint': 'oath/token',
            'pd2':'post_Data3',
            'headers1':code1,	
            'comment': vrn1}

        )    

class VatLiabilitiesView(View):
    def get(self, request):
        tok=mtd_tokens.objects.latest('id')
        obj=apps_ids.objects.get(id=5)
        posted_obj = posted.objects.latest('id') #filter(vrn=tok.vrn)[:0].get()
        periodpj=posted_obj.periodKey
        tokpj=tok.access_token



        # generating access token code
        endpoint1 = "https://test-api.service.hmrc.gov.uk/oauth/token"
        print(tok.refresh_token)
        post_data ={
            'client_secret' :'f18af97f-2acb-41f9-92d2-e373c44618a8',
            'client_id':'ZjScpZHoNA6TQJytfrEkFcSryf4a',
            'grant_type':'refresh_token',
            'refresh_token':tok.refresh_token
        }

        response_vat1 = requests.post(endpoint1, data=post_data)
        print('oath api')
        print(response_vat1)
        oath_api_response = response_vat1.json()
        print(oath_api_response)
        if 'error' in oath_api_response: 
            api_status = {'status_code': oath_api_response['error'], 'reason':oath_api_response['error_description']}
            print('coming in auth error')
            return render(request, 'pages/mtd/subnav/vat_liabilities.html', { 'result': oath_api_response , 'full_path': request.get_full_path(),'vrn':tok.vrn, 'api_status': api_status})
        else:
            print('auth successful')
            endpoint='https://test-api.service.hmrc.gov.uk/organisations/vat/'+tok.vrn +'/liabilities' 

            #endpoint='https://test-api.service.hmrc.gov.uk/organisations/vat/'+tok.vrn +'/obligations/'

            headers ={
                'Authorization': 'Bearer '+ oath_api_response["access_token"],  #'5c5adcd9d4ee25ac173c78225ef45462',
                'Accept' :'application/vnd.hmrc.1.0+json',
                #'Gov-Test-Scenario':'QUARTERLY_NONE_MET',
                #	'Gov-Test-Scenario':'INVALID_PERIODKEY',

                'Content-Type':'application/json'#,
                #"from": "2018-01-01", 
                # "to": "2018-12-31"
            }


            post_Data2={
            'from': '2019-06-14', 
            'to': '2020-05-29',
            #'status':'F'	  
                    }


            l= requests.get(endpoint,data=json.dumps(post_Data2), headers=headers)
            if l.ok:

                print(l.request.body)
                print(l.request.headers)
                print(l.request.url)
                try:
                    lj=l.json()
                except:
                    print('failed json in try/except')
                else:
                    print('PASS json details ' + lj)
                    api_status = {'status_code': l.status_code, 'reason':l.reason}
                    return render(request, 'pages/mtd/subnav/vat_liabilities.html', {'result': lj, 'full_path': request.get_full_path(), 'vrn':tok.vrn, 'api_status': api_status})

            else:
                error = l.json()
                print('failed get l.ok is:' + str(l.ok) + ' status code:' + str(l.status_code) + ' reason code: ' + str(l.reason))
                api_status = {'status_code': error['statusCode'], 'reason':error['message']}
                print ('detailed fail reason: ')
                print( l.json())        
                return render(request, 'pages/mtd/subnav/vat_liabilities.html', { 'result': l , 'full_path': request.get_full_path(),'vrn':tok.vrn, 'api_status': api_status})

class VatPaymentView(View):
    def get(self, request):
        tok=mtd_tokens.objects.latest('id')
        obj=apps_ids.objects.get(id=5)
        posted_obj = posted.objects.latest('id') #filter(vrn=tok.vrn)[:0].get()
        periodpj=posted_obj.periodKey
        tokpj=tok.access_token


        endpoint='https://test-api.service.hmrc.gov.uk/organisations/vat/'+tok.vrn +'/payments'#+ periodpj #"%23001"  

        #endpoint='https://test-api.service.hmrc.gov.uk/organisations/vat/'+tok.vrn +'/obligations/'

        headers ={
            'Authorization': 'Bearer '+ tok.access_token,  #'5c5adcd9d4ee25ac173c78225ef45462',
            'Accept' :'application/vnd.hmrc.1.0+json',
            #'Gov-Test-Scenario':'QUARTERLY_NONE_MET',
            #	'Gov-Test-Scenario':'INVALID_PERIODKEY',

            'Content-Type':'application/json'
        }

        post_Data2={
        #'to': '2019-06-15',

        #'from': '2019-06-12' 
        #'status':'F'	  
                }

        p=[]
        pj=[]		
        p= requests.get(endpoint,data=json.dumps(post_Data2), headers=headers)
                




        print(p.request.body)


        print(p.request.headers)

        print(p.request.url)

        pj=p.json()
        return render(request, 'pages/mtd/subnav/vat_payment.html', {'result': pj, 'full_path': request.get_full_path()})


class VatObligationsView(View):
    def get(self, request):
        tok=mtd_tokens.objects.latest('id')
        obj=apps_ids.objects.get(id=5)
        #posted_obj = posted.objects.filter(vrn=tok.vrn)[:1].get()
        #periodvr=posted_obj.periodKey
        tokvj=tok.access_token

        #endpoint='https://test-api.service.hmrc.gov.uk/organisations/vat/'+tok.vrn +'/returns/'+ period  

        endpoint='https://test-api.service.hmrc.gov.uk/organisations/vat/'+tok.vrn +'/obligations'

        headers ={
            'Authorization': 'Bearer '+ tok.access_token,  #'5c5adcd9d4ee25ac173c78225ef45462',
            'Accept' :'application/vnd.hmrc.1.0+json',
            #'Gov-Test-Scenario':'INVALID_PERIODKEY',
            'Content-Type':'application/json'
        }

        #date optional to removing
        post_Data2={
        "from": "2018-06-14", 
        "to": "2018-06-15",
        "status":"F"	  
                }
        l=[]
                #json.dumps(post_Data2)

        o= requests.get(endpoint,data=json.dumps(post_Data2), headers=headers)
                


        print ("print(o.request.body)")
        print(o.request.body)

        print(o.request.headers)

        print(o.request.url)
        print(o.json())

        try:
            print('coming in try')
            api_status = {'status_code': o.status_code, 'reason':o.reason}
            oj=o.json()
        except:
            print('coming in except')
            api_status = {'status_code': o.status_code, 'reason':o.reason}

        return render(request, 'pages/mtd/subnav/vat_obligations.html', {'result': oj, 'full_path': request.get_full_path(), 'vrn':tok.vrn, 'api_status': api_status})

class VatReturnView(View):
    def get(self, request):
        tok=mtd_tokens.objects.latest('id')
        obj=apps_ids.objects.get(id=5)
        posted_obj = posted.objects.latest('id') #filter(vrn=tok.vrn)[:0].get()
        periodpj=posted_obj.periodKey
        tokpj=tok.access_token


        endpoint='https://test-api.service.hmrc.gov.uk/organisations/vat/'+tok.vrn +'/returns/'+ periodpj #"%23001"  

        #endpoint='https://test-api.service.hmrc.gov.uk/organisations/vat/'+tok.vrn +'/obligations/'

        headers ={
            'Authorization': 'Bearer '+ tok.access_token,  #'5c5adcd9d4ee25ac173c78225ef45462',
            'Accept' :'application/vnd.hmrc.1.0+json',
            #'Gov-Test-Scenario':'QUARTERLY_NONE_MET',
            #	'Gov-Test-Scenario':'INVALID_PERIODKEY',

            'Content-Type':'application/json'
        }


        post_Data2={
        'from': '2012006-30', 
        'to': '2020-05-31',
        #'status':'F'	  
                }

        r= requests.get(endpoint,data=json.dumps(post_Data2), headers=headers)
                


        print(r.request.body)


        print(r.request.headers)

        print(r.request.url)
        try:
            print('coming in try')
            print(str(r.status_code) +r.reason)
            api_status = {'status_code': r.status_code, 'reason':r.reason}
            rj=r.json()
        except:
            print('coming in except')
            print(str(r.status_code) +r.reason)
            api_status = {'status_code': r.status_code, 'reason':r.reason}

        return render(request, 'pages/mtd/subnav/vat_return.html', {'result': rj, 'full_path': request.get_full_path(), 'vrn':tok.vrn, 'api_status': api_status}) 

class InvoiceListView(View):
    def get(self, request):
        invoices = invoice.objects.all()
        return render(request, 'pages/mtd/invoice-list.html', { "invoices" : invoices })

class InvoiceCreateView(View):
    def get(self, request):
        return render(request, 'pages/mtd/invoice.html')

class InvoiceView(View):
    def get(self, request, id=None):
        if id:
            invoice_data = None
            invoice_data = invoice.objects.filter(id=id).values()
            if invoice_data:
                invoice_data = invoice_data[0]

                product_data = products.objects.filter(invoice_id = invoice_data['id']).values() if products.objects.filter(invoice_id = invoice_data['id']).values() else None
                
                payment_status_data = payment_status.objects.filter(invoice_id = invoice_data['id']).values() if payment_status.objects.filter(invoice_id = invoice_data['id']).values() else None

                return render(request, 'pages/mtd/invoice.html', {"invoice": invoice_data, "products": product_data, "payment_statuses" : payment_status_data, "id": id})
            else:
                return render(request, 'pages/mtd/invoice.html', {"invoice": None, "products": None, "payment_statuses" : None, "id": None})
        else:
            return render(request, 'pages/mtd/invoice.html', {"invoice": None, "products": None, "payment_statuses" : None, "id": None})

    def post(self, request, id=None):
        check_invoice_data = None
        page_invoice_id = int(request.POST.get('page_invoice_id')) if request.POST.get('page_invoice_id') else None
        
        if page_invoice_id:
            # check_invoice_data variable decides whether it is an edit page or new page
            check_invoice_data = invoice.objects.filter(id=page_invoice_id).values()

        
        

        invoice_ref_num = request.POST.get('invoice_ref_num').strip() if request.POST.get('invoice_ref_num') else None
        date = request.POST.get('date').strip() if request.POST.get('date') else None
        reference = request.POST.get('reference').strip() if request.POST.get('reference') else None
        account = request.POST.get('account').strip() if request.POST.get('account') else None
        client_name = request.POST.get('client_name').strip() if request.POST.get('client_name') else None
        address = request.POST.get('address').strip() if request.POST.get('address') else None
        products_count = request.POST.get('products_sections_count')
        invoice_version = request.POST.get('invoice_version').strip() if request.POST.get('invoice_version') else None
        payment_status_count = request.POST.get('payment_status_count')
        self.process_invoice_data(request, invoice_ref_num, date, reference, account, client_name, address, products_count, invoice_version, check_invoice_data, page_invoice_id, payment_status_count)
        return redirect('/mtd/invoice-list/')

    def process_invoice_data(self, request, invoice_ref_num, date, reference, account, client_name, address, products_count, invoice_version, check_invoice_data, page_invoice_id, payment_status_count):
        save_query = None
        # if check_invoice_data does not exist then save data into invoice table otherwise update invoice table
        if check_invoice_data is None:
            print('saving')
            save_query = invoice(invoice_ref_num=invoice_ref_num, date = date, reference = reference, account = account, client_name = client_name, address = address, invoice_version = invoice_version)
            save_query.save()
            print('save_query.id')
            print(save_query.id)
            saved_id = save_query.id  # latest saved record it
        else:
            # update invoice table corresponding page_invoice_id
            print('updating')
            save_query = invoice.objects.filter(id=page_invoice_id).update(invoice_ref_num=invoice_ref_num, date = date, reference = reference, account = account, client_name = client_name, address = address, invoice_version = invoice_version)
            saved_id = page_invoice_id       # latest updated record id

        # process products
        products_result = self.process_products_data(request, products_count, saved_id, check_invoice_data)

        # process payment statuses
        self.process_payment_statuses_data(request, payment_status_count, saved_id)

        # update existing invoice with gross, net, vat attributes returned from process_products_data() method
        invoice.objects.filter(id = saved_id).update(net = round(products_result["net"],2), gross = round(products_result["gross"], 2), vat = round(products_result["vat"], 2))
        # invoice.objects.filter(id = saved_id).update(net = products_result["net"], gross = products_result["gross"], vat = products_result["vat"])
        

    def process_products_data(self, request, products_count, latest_id, check_invoice_data):
        price_list = []
        gross_list = []
        price_default = 0
        gross_default = 0
        all_product_ids = []
        
        for single_count in range(int(products_count)):
            
            product_hidden_id = request.POST.get('product_hidden_id_'+str(single_count)) if request.POST.get('product_hidden_id_'+str(single_count)) else None
            description = request.POST.get('description_'+str(single_count)).strip() if request.POST.get('description_'+str(single_count)) else None
            # if product deleted from DOM, then we need to continue loop for next index instead of moving further.
            if description is None:
                continue
            unit_price  = float(request.POST.get('unit_price_'+str(single_count)).strip()) if request.POST.get('unit_price_'+str(single_count)) else None
            quantity =  float(request.POST.get('quantity_'+str(single_count)).strip()) if request.POST.get('quantity_'+str(single_count)) else None
            vat_rate =  float(request.POST.get('vat_rate_'+str(single_count)).strip()) if request.POST.get('vat_rate_'+str(single_count)) else None
            if description:
                price = quantity * unit_price   # price excluding vat
                price_list.append(price)
                gross = price + (price * vat_rate)/100
                gross_list.append(gross)
                if product_hidden_id:
                    # update product as product id already exist.
                    products.objects.filter(id=product_hidden_id).update(invoice_id=latest_id, description=description, unit_price=str(unit_price), quantity=str(quantity), vat_rate = str(vat_rate))
                    all_product_ids.append(int(product_hidden_id))
                    print('product update')
                    print(product_hidden_id)
                else:
                    # save product as product id does not exist.
                    save_products = products(invoice_id=latest_id, description=description, unit_price=str(unit_price), quantity=str(quantity), vat_rate = str(vat_rate))
                    save_products.save()

                    all_product_ids.append(save_products.id)
                    print('product add')
                    print(save_products.id)

                print(price)
                print(gross)
                

        # delete records deleted from DOM
        db_products = list(products.objects.filter(invoice_id=latest_id).values())    # db products
        for db_product in db_products:
            if db_product['id'] not in all_product_ids:
                products.objects.filter(id=db_product['id']).delete()
                
            

        # get total of prices
        if price_list:
            for pr in price_list:
                price_default += pr

        # get total of gross values
        if gross_list:
            for gr in gross_list:
                gross_default += gr

        print('total')
        print(price_default)
        print(gross_default)
        net = price_default     # total of all products prices
        gross = gross_default   # total of all products gross prices
        vat = gross - net 
        print('vat %s'% vat)
        return {"net": net, "gross": gross, "vat" : vat}


    def process_payment_statuses_data(self, request, payment_status_count, saved_id):
        all_payment_status_ids = []
        for single_count in range(int(payment_status_count)):
            payment_status_hidden_id = int(request.POST.get('payment_status_id_'+str(single_count))) if request.POST.get('payment_status_id_'+str(single_count)) else None 
            payment_status_value = request.POST.get('payment_status_'+str(single_count)) if request.POST.get('payment_status_'+str(single_count)) else None 

            # if product deleted from DOM, then we need to continue loop for next index instead of moving further.
            if payment_status_value is None:
                continue

            # update payment_status table if  payment_status_hidden_id variable exists otherwise add new record into payment_status table.
            if payment_status_hidden_id:
                payment_status.objects.filter(id=payment_status_hidden_id).update(invoice_id = saved_id, status = payment_status_value)
                all_payment_status_ids.append(payment_status_hidden_id)
            else:
                # print('%s %s'% (saved_id, payment_status))
                save_payment_status = payment_status(invoice_id = saved_id, status = payment_status_value)
                save_payment_status.save()
                print('saved id')
                print(save_payment_status.id)
                all_payment_status_ids.append(save_payment_status.id)

        # delete payment status record deleted from DOM
        print('all_payment_status_ids')
        print(all_payment_status_ids)
        db_payment_statuses = list(payment_status.objects.filter(invoice_id = saved_id).values())
        print(db_payment_statuses)
        if db_payment_statuses:
            for single in db_payment_statuses:
                if single['id'] not in all_payment_status_ids:
                    payment_status.objects.filter(id = single['id']).delete()

class InvoiceDetailsView(View):
    def get(self, request, id):
        invoice_details = None
        products_details = []
        payment_statuses_details = []
        invoice_check = list(invoice.objects.filter(id=id).values())
        if invoice_check:
            invoice_details = invoice_check[0]
            products_details = list(products.objects.filter(invoice_id=invoice_details['id']).values())
            payment_statuses_details = list(payment_status.objects.filter(invoice_id=invoice_details['id']).values())
        return render(request, 'pages/mtd/invoice_details.html', {"invoice": invoice_details, "products":products_details, "payment_statuses" : payment_statuses_details})